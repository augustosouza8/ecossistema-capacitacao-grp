from pathlib import Path
from typing import TYPE_CHECKING, cast

import openpyxl

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

CATALOG_PATH = Path(__file__).parent.parent / "data" / "imports" / "materials.xlsx"


def import_xlsx() -> None:
    if not CATALOG_PATH.exists():
        print(f"Arquivo XLSX nao encontrado: {CATALOG_PATH}")  # noqa: T201
        return

    workbook = openpyxl.load_workbook(CATALOG_PATH, read_only=True, data_only=True)
    try:
        worksheet = cast("Worksheet", workbook.active)
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None
        )
        if header_row is None:
            print("Planilha vazia.")  # noqa: T201
            return

        count = sum(
            1
            for row in worksheet.iter_rows(min_row=2, values_only=True)
            if row and row[0] is not None
        )
    finally:
        workbook.close()

    print("Catalogo Excel validado com sucesso.")  # noqa: T201
    print(f"Arquivo: {CATALOG_PATH}")  # noqa: T201
    print(f"Colunas encontradas: {', '.join(str(value) for value in header_row)}")  # noqa: T201
    print(f"Total de materiais ativos na fonte: {count}")  # noqa: T201


if __name__ == "__main__":
    import_xlsx()
