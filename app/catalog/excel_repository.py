from collections.abc import Mapping
from pathlib import Path

from openpyxl import load_workbook

from app.catalog.models import CatalogMaterial
from app.catalog.repository import CatalogRepository

REQUIRED_HEADERS = {
    "id",
    "type",
    "title",
    "module",
    "theme",
    "subtheme",
    "subsubtheme",
    "keywords",
    "summary",
    "source_url",
    "blob_path",
}


class ExcelCatalogRepository(CatalogRepository):
    def __init__(self, workbook_path: Path) -> None:
        self._workbook_path = workbook_path
        self._cached_materials: tuple[CatalogMaterial, ...] = ()
        self._cached_mtime_ns: int | None = None
        self._cached_metadata: Mapping[str, object] = {
            "path": str(workbook_path),
            "headers": (),
            "row_count": 0,
        }

    def list_materials(self) -> list[CatalogMaterial]:
        self._ensure_cache_loaded()

        return list(self._cached_materials)

    def get_metadata(self) -> Mapping[str, object]:
        self._ensure_cache_loaded()
        return self._cached_metadata

    def _ensure_cache_loaded(self) -> None:
        current_mtime_ns = self._workbook_path.stat().st_mtime_ns
        if current_mtime_ns == self._cached_mtime_ns:
            return

        materials, metadata = self._load_materials()
        self._cached_materials = tuple(materials)
        self._cached_metadata = metadata
        self._cached_mtime_ns = current_mtime_ns

    def _load_materials(self) -> tuple[list[CatalogMaterial], Mapping[str, object]]:
        workbook = load_workbook(self._workbook_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            if worksheet is None:
                raise ValueError("A planilha do catalogo nao pode ser carregada.")

            header_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                None,
            )
            if header_row is None:
                raise ValueError("A planilha do catalogo esta vazia.")

            headers = [self._normalize_header(value) for value in header_row]
            self._validate_headers(headers)

            materials: list[CatalogMaterial] = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue

                row_data = dict[str, object](zip(headers, row, strict=False))
                materials.append(self._build_material(row_data))

            active_materials = sorted(
                (material for material in materials if material.is_active),
                key=lambda material: material.id,
            )
            metadata: Mapping[str, object] = {
                "path": str(self._workbook_path),
                "headers": tuple(headers),
                "row_count": len(active_materials),
            }
            return active_materials, metadata
        finally:
            workbook.close()

    def _validate_headers(self, headers: list[str]) -> None:
        missing_headers = REQUIRED_HEADERS.difference(headers)
        if missing_headers:
            missing = ", ".join(sorted(missing_headers))
            raise ValueError(
                f"A planilha do catalogo nao possui as colunas obrigatorias: {missing}."
            )

    def _build_material(self, row_data: Mapping[str, object]) -> CatalogMaterial:
        return CatalogMaterial(
            id=self._parse_int(row_data["id"]),
            type=self._normalize_required_text(row_data["type"]),
            title=self._normalize_required_text(row_data["title"]),
            module=self._normalize_required_text(row_data["module"]),
            theme=self._normalize_optional_text(row_data.get("theme")),
            subtheme=self._normalize_optional_text(row_data.get("subtheme")),
            subsubtheme=self._normalize_optional_text(row_data.get("subsubtheme")),
            keywords=self._normalize_optional_text(row_data.get("keywords")),
            summary=self._normalize_optional_text(row_data.get("summary")),
            source_url=self._normalize_optional_text(row_data.get("source_url")),
            blob_path=self._normalize_optional_text(row_data.get("blob_path")),
            is_active=self._parse_bool(row_data.get("is_active")),
        )

    @staticmethod
    def _normalize_header(value: object) -> str:
        if value is None:
            return ""

        return str(value).strip()

    @staticmethod
    def _normalize_required_text(value: object) -> str:
        text = ExcelCatalogRepository._normalize_optional_text(value)
        if text is None:
            raise ValueError(
                "A planilha do catalogo possui um campo obrigatorio vazio."
            )

        return text

    @staticmethod
    def _normalize_optional_text(value: object) -> str | None:
        if value is None:
            return None

        text = str(value).strip()
        return text or None

    @staticmethod
    def _parse_int(value: object) -> int:
        if value is None:
            raise ValueError("A planilha do catalogo possui um ID vazio.")

        if isinstance(value, int):
            return value

        if isinstance(value, float):
            return int(value)

        if isinstance(value, str):
            return int(value)

        raise ValueError("A planilha do catalogo possui um ID invalido.")

    @staticmethod
    def _parse_bool(value: object) -> bool:
        if value is None:
            return True

        if isinstance(value, bool):
            return value

        normalized_value = str(value).strip().casefold()
        return normalized_value not in {"0", "false", "nao", "no", "off"}
